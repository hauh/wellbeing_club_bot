"""Get posts from Excel file."""

from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from bot import FILE_TIME


class ParseError(Exception):
	"""Indicates error in Excel file."""


def parse_document(file):
	try:
		document = load_workbook(file)
	except (OSError, IOError, KeyError, InvalidFileException) as e:
		raise ParseError("Ошибка чтения файла.") from e

	images = {}
	# pylint: disable=protected-access
	for image in document.active._images:
		row = image.anchor._from.row + 1
		if row in images:
			raise ParseError(f"Несколько изображений в строке {row}.")
		image.ref.seek(0)
		images[row] = image.ref

	now = datetime.now(FILE_TIME)
	posts = []
	for row in document.active.iter_rows(min_row=2, min_col=2, max_col=6):
		date_cell, time_cell, title_cell, body_cell, points_cell = tuple(row)
		if not date_cell.value:
			continue

		try:
			post_time = datetime.combine(date_cell.value, time_cell.value, FILE_TIME)
		except TypeError as e:
			raise ParseError(
				f"Неправильный формат даты или времени в строке {date_cell.row}."
			) from e
		if post_time < now:
			continue

		if not title_cell.value:
			raise ParseError(f"В ячейке {title_cell.coordinate} должен быть заголовок.")
		post_text = f"*{title_cell.value}*"
		if body_cell.value:
			post_text += "\n\n" + str(body_cell.value)
		if points_cell.value:
			post_text += "\n\n" + "\n".join(
				"✔️ " + line for line in str(points_cell.value).splitlines() if line)

		posts.append([post_time, post_text, images.get(date_cell.row)])

	document.close()
	return posts
